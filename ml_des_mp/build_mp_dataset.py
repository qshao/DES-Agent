"""Assemble the QSPR melting-point training set.

Merges two sources, keyed by InChIKey:
  1. The Jean-Claude Bradley open (curated, multi-source) melting-point set,
     downloaded from figshare. Labels are experimental, multi-source averages.
  2. The in-domain DES pure-component melting points from the project's own
     training CSV, which take precedence so the QSPR stays consistent with the
     distribution the eutectic physics model was trained on.

Writes a committed CSV (``smiles,inchikey,tm_k,source``) used by
``ml_des_mp.train_mp_qspr``.

Usage:
    python -m ml_des_mp.build_mp_dataset
"""
from __future__ import annotations

import argparse
import csv
import urllib.request
from collections import defaultdict
from pathlib import Path

from rdkit import Chem

BRADLEY_URL = "https://ndownloader.figshare.com/files/1503991"
T_MIN_K, T_MAX_K = 100.0, 700.0


def _download(url: str, dest: Path) -> Path:
    if dest.exists():
        return dest
    dest.parent.mkdir(parents=True, exist_ok=True)
    req = urllib.request.Request(url, headers={"User-Agent": "Mozilla/5.0"})
    dest.write_bytes(urllib.request.urlopen(req).read())
    return dest


def _clean(smiles: str):
    mol = Chem.MolFromSmiles((smiles or "").strip())
    if mol is None:
        return None
    return Chem.MolToSmiles(mol), Chem.MolToInchiKey(mol)


def _bradley_rows(xlsx_path: Path):
    import openpyxl

    wb = openpyxl.load_workbook(xlsx_path, read_only=True)
    ws = wb[wb.sheetnames[0]]
    it = ws.iter_rows(values_only=True)
    header = [str(h) for h in next(it)]
    si, mi = header.index("smiles"), header.index("mpC")
    for row in it:
        smiles, mpc = row[si], row[mi]
        if smiles is None or mpc is None:
            continue
        try:
            tm_k = float(mpc) + 273.15
        except (TypeError, ValueError):
            continue
        yield smiles, tm_k


def _des_rows(csv_path: Path):
    for r in csv.DictReader(open(csv_path, encoding="utf-8")):
        for s_col, t_col in (("Smiles#1", "T#1"), ("Smiles#2", "T#2")):
            sm, tv = (r.get(s_col) or "").strip(), (r.get(t_col) or "").strip()
            if not sm or not tv:
                continue
            try:
                yield sm, float(tv)
            except ValueError:
                continue


def build(out_path: Path, des_csv: Path, cache: Path) -> dict:
    xlsx = _download(BRADLEY_URL, cache)
    # InChIKey -> (canonical_smiles, [tm_k...], source)
    bradley: dict[str, tuple[str, list[float]]] = defaultdict(lambda: ("", []))
    for smiles, tm_k in _bradley_rows(xlsx):
        if not (T_MIN_K <= tm_k <= T_MAX_K):
            continue
        cleaned = _clean(smiles)
        if cleaned is None:
            continue
        canon, key = cleaned
        bradley[key] = (canon, bradley[key][1] + [tm_k])

    des: dict[str, tuple[str, list[float]]] = defaultdict(lambda: ("", []))
    for smiles, tm_k in _des_rows(des_csv):
        if not (T_MIN_K <= tm_k <= T_MAX_K):
            continue
        cleaned = _clean(smiles)
        if cleaned is None:
            continue
        canon, key = cleaned
        des[key] = (canon, des[key][1] + [tm_k])

    rows = []
    keys = set(bradley) | set(des)
    n_des = 0
    for key in keys:
        if key in des:  # in-domain DES values win
            canon, vals = des[key]
            source = "des"
            n_des += 1
        else:
            canon, vals = bradley[key]
            source = "bradley"
        tm = sum(vals) / len(vals)
        rows.append((canon, key, round(tm, 3), source))

    rows.sort(key=lambda r: r[1])
    out_path.parent.mkdir(parents=True, exist_ok=True)
    with open(out_path, "w", newline="", encoding="utf-8") as f:
        w = csv.writer(f)
        w.writerow(["smiles", "inchikey", "tm_k", "source"])
        w.writerows(rows)
    return {"total": len(rows), "des": n_des, "bradley": len(rows) - n_des}


def main(argv=None) -> None:
    ap = argparse.ArgumentParser()
    ap.add_argument("--out", default="ml_des_mp/mp_train.csv")
    ap.add_argument("--des-csv", default="ml_des_mp/Melting_temperature_appended_35il_03082026.csv")
    ap.add_argument("--cache", default="ml_des_mp/.cache/bradley_full.xlsx")
    args = ap.parse_args(argv)
    stats = build(Path(args.out), Path(args.des_csv), Path(args.cache))
    print(f"Wrote {stats['total']} rows to {args.out} "
          f"(des={stats['des']}, bradley={stats['bradley']})")


if __name__ == "__main__":
    main()
